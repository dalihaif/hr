"""
Excel导入导出工具类
使用openpyxl库处理Excel文件
"""
import io
from openpyxl import Workbook
from openpyxl import load_workbook


def export_employees_to_excel(employees):
    """
    导出职工信息到Excel
    :param employees: 职工列表
    :return: Excel文件的字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "职工花名册"
    
    # 表头
    headers = ['工号', '姓名', '性别', '出生日期', '民族', '政治面貌', 
               '婚姻状况', '籍贯', '联系电话', '邮箱', '状态', '入职日期',
               '当前科室', '当前岗位', '当前职称']
    ws.append(headers)
    
    # 数据行
    for emp in employees:
        row = [
            emp.get('emp_no', ''),
            emp.get('name', ''),
            emp.get('gender', ''),
            emp.get('birth_date', ''),
            emp.get('ethnicity', ''),
            emp.get('political_status', ''),
            emp.get('marital_status', ''),
            emp.get('native_place', ''),
            emp.get('phone_display', emp.get('phone', '')),
            emp.get('email', ''),
            emp.get('status', ''),
            emp.get('entry_date', ''),
            emp.get('current_dept', ''),
            emp.get('current_position', ''),
            emp.get('current_title', '')
        ]
        ws.append(row)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_employees_from_excel(file_stream):
    """
    从Excel导入职工信息
    :param file_stream: Excel文件流
    :return: 职工数据列表
    """
    wb = load_workbook(file_stream)
    ws = wb.active
    
    employees = []
    # 跳过表头，从第二行开始
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 跳过空行
            continue
        
        emp = {
            'emp_no': str(row[0]) if row[0] else '',
            'name': str(row[1]) if row[1] else '',
            'gender': str(row[2]) if row[2] else '',
            'birth_date': str(row[3]) if row[3] else '',
            'ethnicity': str(row[4]) if row[4] else '汉',
            'political_status': str(row[5]) if row[5] else '群众',
            'marital_status': str(row[6]) if row[6] else '未婚',
            'native_place': str(row[7]) if row[7] else '',
            'phone': str(row[8]) if row[8] else '',
            'email': str(row[9]) if row[9] else '',
            'entry_date': str(row[11]) if row[11] else ''
        }
        
        # 验证必填字段
        if emp['emp_no'] and emp['name']:
            employees.append(emp)
    
    return employees


def export_salary_to_excel(salary_records):
    """
    导出工资记录到Excel
    :param salary_records: 工资记录列表
    :return: Excel文件的字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "工资表"
    
    # 表头
    headers = ['月份', '工号', '姓名', '科室', '基本工资', '岗位津贴', 
               '医疗津贴', '住房补贴', '夜班费', '加班费', '应发合计',
               '养老保险', '医疗保险', '失业保险', '住房公积金', '社保扣缴合计',
               '个人所得税', '其他扣款', '扣款合计', '实发工资', '状态']
    ws.append(headers)
    
    # 数据行
    for record in salary_records:
        row = [
            record.get('month', ''),
            record.get('emp_no', ''),
            record.get('name', ''),
            record.get('dept', ''),
            record.get('base_salary', 0),
            record.get('position_allowance', 0),
            record.get('medical_allowance', 0),
            record.get('housing_allowance', 0),
            record.get('night_shift_pay', 0),
            record.get('overtime_pay', 0),
            record.get('gross_salary', 0),
            record.get('pension_personal', 0),
            record.get('medical_personal', 0),
            record.get('unemployment_personal', 0),
            record.get('housing_fund_personal', 0),
            record.get('total_deduction_social', 0),
            record.get('tax_deduction', 0),
            record.get('other_deduction', 0),
            record.get('total_deduction', 0),
            record.get('net_salary', 0),
            record.get('status', '')
        ]
        ws.append(row)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_performance_to_excel(assessments):
    """
    导出绩效结果到Excel
    :param assessments: 考核记录列表
    :return: Excel文件的字节流
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "绩效考核结果"
    
    # 表头
    headers = ['周期', '工号', '姓名', '科室', '考核模板', '总分', '最终得分', '等级', '状态']
    ws.append(headers)
    
    # 数据行
    for assess in assessments:
        row = [
            assess.get('period', ''),
            assess.get('emp_no', ''),
            assess.get('name', ''),
            assess.get('dept', ''),
            assess.get('template_name', ''),
            assess.get('total_score', ''),
            assess.get('final_score', ''),
            assess.get('level', ''),
            assess.get('status', '')
        ]
        ws.append(row)
    
    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
